import datetime
import sys

from utils import clean_id


def load_xls(fn):
    'Parse .xls file at `fn`.  Yield one row per data row in the sheet.'

    import xlrd

    def convert_cell(c):
        'Return properly-typed value for xls Cell `c`.'
        if c.ctype == xlrd.XL_CELL_NUMBER:
            v = int(c.value)
            if v != c.value:
                return c.value
            return v

        if c.ctype == xlrd.XL_CELL_DATE:
            t = xlrd.xldate.xldate_as_tuple(c.value, 0)
            dt = datetime.datetime(*t)
            return dt.isoformat(sep=' ')

        if c.ctype == xlrd.XL_CELL_BOOLEAN:
            return bool(c.value)

        if c.ctype == xlrd.XL_CELL_ERROR:
            print(f'error: {c.value}', file=sys.stderr)
            return f'error: {c.value}'

        v = c.value.strip()
        if v:
            return v

    wb = xlrd.open_workbook(fn, logfile=sys.stderr)

    for sheet in wb:
        hdrs = [clean_id(convert_cell(c)).lower() for c in sheet.row(0)]

        for i in range(1, sheet.nrows):
            row = {}
            for k, v in zip(hdrs, [convert_cell(c) for c in sheet.row(i)]):
                if v is not None:
                    row[k] = v
            if row:
                yield sheet.name, row


def load_xlsx(fn):
    'Parse .xlsx file at `fn`.  Yield one row per data row in the sheet.'

    import openpyxl

    def convert_cell(c):
        if c.value is None:
            return None
        if isinstance(c.value, str):
            v = c.value.strip()
            if v:
                return v
        return c.value

    wb = openpyxl.load_workbook(fn, data_only=True, read_only=True)

    for sheet in wb.worksheets:
        if sheet.max_row <= 1:
            continue

        tblname = clean_id(sheet.title)

        rows = sheet.iter_rows()
        hdrs = [clean_id(convert_cell(c)).lower() for c in next(rows) if c.value is not None]

        for row in rows:
            row = {}
            for k, v in zip(hdrs, [convert_cell(c) for c in row]):
                if v is not None:
                    row[k] = v

            if row:
                yield tblname, row
